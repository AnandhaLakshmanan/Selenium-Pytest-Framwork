import openpyxl


class HomePageTestData:
    test_home_page_data = [{"name": "john doe", "email": "company@gmail.com", "gender": "Male"},
                           {"name": "jane doe", "email": "company2@gmail.com", "gender": "Female"}]

    @staticmethod
    def get_test_data_from_excel(test_case_name):
        book = openpyxl.load_workbook("test_data/test_data.xlsx")
        sheet = book.active
        test_data = []
        for row in range(1, sheet.max_row + 1):
            data = {}
            if sheet.cell(row=row, column=1).value == test_case_name:
                for col in range(2, sheet.max_column + 1):
                    data[sheet.cell(row=1, column=col).value] = sheet.cell(row=row, column=col).value
                test_data.append(data)
        return test_data
